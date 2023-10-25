'''
Autor: Luiz Fernando Antonelli Galati
Julho/2023
'''

'''
Este código lê um documento no formato .xlsx, realiza análises estatísticas dos dados presentes no documento e escreve os
resultados no próprio documento.
'''

import openpyxl
import xlsxwriter

def escreveCabecalho (planilha):
    planilha.write (0, 1, "1º semestre")
    planilha.write (0, 3, "3º semestre")
    planilha.write (0, 5, "5º semestre")
    planilha.write (0, 7, "7º semestre")
    planilha.write (0, 9, "9º semestre")

    planilha.write (1, 0, "Objetivos do programa de graduação FGV Direito SP")
    planilha.write (2, 0, "Domínio de conceitos, estruturas e racionalidades fundamentais do Direito")
    planilha.write (3, 0, "Conhecimento de áreas contíguas ao Direito")
    planilha.write (4, 0, "Aplicação prática de conceitos e estruturas do Direito")
    planilha.write (5, 0, "Pesquisa Jurídica")
    planilha.write (6, 0, "Comunicação")
    planilha.write (7, 0, "Colaboração e trabalho em rede")
    planilha.write (8, 0, "Ética")
    planilha.write (9, 0, "Empreendedorismo")
    planilha.write (10, 0, "Cosmopolitanismo")

    planilha.write (12, 0, "Total de disciplinas")
    planilha.write (13, 0, "Total de disciplinas avaliadas")
    planilha.write (14, 0, "Pontuação máxima para cada quesito")

    i = 1
    while (i < 10):    	
        planilha.write (1, i, "Contribuição total das disciplinas")
        planilha.write (1, i + 1, "Porcentagem da pontuação máxima")
        i = i + 2

    planilha.write (0, 11, "Todos os semestres")
    planilha.write (1, 11, "Contribuição total das disciplinas")
    planilha.write (1, 12, "Porcentagem da pontuação máxima")


def main ():
    arquivo = openpyxl.load_workbook ("(2022) Dados por disciplina.xlsx")
    numPlanilhas = len (arquivo.sheetnames)

    novoArquivo = xlsxwriter.Workbook ("(2022) Dados agregados-novo.xlsx")
    novaPlanilha = novoArquivo.add_worksheet ("(2022) Dados agregados")

    escreveCabecalho (novaPlanilha)

    somaDominioConceitos = 0
    somaConhecimento = 0
    somaAplicacao = 0
    somaPesquisa = 0
    somaComunicacao = 0
    somaColaboracao = 0
    somaEtica = 0
    somaEmpreendedorismo = 0
    somaCosmopolitanismo = 0

    somaTotalDisciplinas = 0
    somaDisciplinasAvaliadas = 0

    i = 0
    b = 1
    while (i < len (arquivo.sheetnames)):
        nomePlanilha = arquivo.sheetnames[i]
        planilha = arquivo[nomePlanilha]        
        
        dominioConceitos = 0       
        conhecimento = 0       
        aplicacao = 0        
        pesquisa = 0
        comunicacao = 0        
        colaboracao = 0        
        etica = 0        
        empreendedorismo = 0        
        cosmopolitanismo = 0                

        totalDisciplinas = 0
        disciplinasAvaliadas = 0

        j = 3
        while (j <= planilha.max_column):
            if (planilha.cell(2, j - 1).value == "Objetivos da disciplina"):
                totalDisciplinas = totalDisciplinas + 1

            grauContribuicao = planilha.cell(3, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):
                disciplinasAvaliadas = disciplinasAvaliadas + 1                                                  
                dominioConceitos = dominioConceitos + grauContribuicao

            grauContribuicao = planilha.cell(4, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                conhecimento = conhecimento + grauContribuicao

            grauContribuicao = planilha.cell(5, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                aplicacao = aplicacao + grauContribuicao

            grauContribuicao = planilha.cell(6, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                pesquisa = pesquisa + grauContribuicao

            grauContribuicao = planilha.cell(7, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                comunicacao = comunicacao + grauContribuicao

            grauContribuicao = planilha.cell(8, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                colaboracao = colaboracao + grauContribuicao

            grauContribuicao = planilha.cell(9, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                etica = etica + grauContribuicao

            grauContribuicao = planilha.cell(10, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                empreendedorismo = empreendedorismo + grauContribuicao

            grauContribuicao = planilha.cell(11, j).value                                       
            if (grauContribuicao == 0 or grauContribuicao == 1 or grauContribuicao == 2 or grauContribuicao == 3):                                                    
                cosmopolitanismo = cosmopolitanismo + grauContribuicao

            j = j + 2

        pontMax = disciplinasAvaliadas*3

        novaPlanilha.write (2, b, dominioConceitos)
        novaPlanilha.write (2, b + 1, dominioConceitos/pontMax)      
        somaDominioConceitos = somaDominioConceitos + dominioConceitos

        novaPlanilha.write (3, b, conhecimento)
        novaPlanilha.write (3, b + 1, conhecimento/pontMax)
        somaConhecimento = somaConhecimento + conhecimento        

        novaPlanilha.write (4, b, aplicacao)
        novaPlanilha.write (4, b + 1, aplicacao/pontMax)
        somaAplicacao = somaAplicacao + aplicacao

        novaPlanilha.write (5, b, pesquisa)
        novaPlanilha.write (5, b + 1, pesquisa/pontMax)
        somaPesquisa = somaPesquisa + pesquisa

        novaPlanilha.write (6, b, comunicacao)
        novaPlanilha.write (6, b + 1, comunicacao/pontMax)
        somaComunicacao = somaComunicacao + comunicacao
        
        novaPlanilha.write (7, b, colaboracao)
        novaPlanilha.write (7, b + 1, colaboracao/pontMax)
        somaColaboracao = somaColaboracao + colaboracao

        novaPlanilha.write (8, b, etica)
        novaPlanilha.write (8, b + 1, etica/pontMax)
        somaEtica = somaEtica + etica

        novaPlanilha.write (9, b, empreendedorismo)
        novaPlanilha.write (9, b + 1, empreendedorismo/pontMax)
        somaEmpreendedorismo = somaEmpreendedorismo + empreendedorismo

        novaPlanilha.write (10, b, cosmopolitanismo)
        novaPlanilha.write (10, b + 1, cosmopolitanismo/pontMax)
        somaCosmopolitanismo = somaCosmopolitanismo + cosmopolitanismo


        novaPlanilha.write (12, b, totalDisciplinas)        
        somaTotalDisciplinas = somaTotalDisciplinas + totalDisciplinas

        novaPlanilha.write (13, b, disciplinasAvaliadas)
        somaDisciplinasAvaliadas = somaDisciplinasAvaliadas + disciplinasAvaliadas

        novaPlanilha.write (14, b, pontMax)

        b = b + 2
        i = i + 1

    pontMaxTotal = somaDisciplinasAvaliadas*3    

    novaPlanilha.write (2, 11, somaDominioConceitos)
    novaPlanilha.write (2, 12, somaDominioConceitos/pontMaxTotal)
    novaPlanilha.write (3, 11, somaConhecimento)
    novaPlanilha.write (3, 12, somaConhecimento/pontMaxTotal)
    novaPlanilha.write (4, 11, somaAplicacao)
    novaPlanilha.write (4, 12, somaAplicacao/pontMaxTotal)
    novaPlanilha.write (5, 11, somaPesquisa)
    novaPlanilha.write (5, 12, somaPesquisa/pontMaxTotal)
    novaPlanilha.write (6, 11, somaComunicacao)
    novaPlanilha.write (6, 12, somaComunicacao/pontMaxTotal)
    novaPlanilha.write (7, 11, somaColaboracao)
    novaPlanilha.write (7, 12, somaColaboracao/pontMaxTotal)
    novaPlanilha.write (8, 11, somaEtica)    
    novaPlanilha.write (8, 12, somaEtica/pontMaxTotal)
    novaPlanilha.write (9, 11, somaEmpreendedorismo)
    novaPlanilha.write (9, 12, somaEmpreendedorismo/pontMaxTotal)
    novaPlanilha.write (10, 11, somaCosmopolitanismo)
    novaPlanilha.write (10, 12, somaCosmopolitanismo/pontMaxTotal)
    
    novaPlanilha.write (12, 11, somaTotalDisciplinas)
    novaPlanilha.write (13, 11, somaDisciplinasAvaliadas)    
    novaPlanilha.write (14, 11, pontMaxTotal)

    novoArquivo.close ()

main ()
